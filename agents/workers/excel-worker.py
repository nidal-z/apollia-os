"""excel-worker — Worker Agent for Excel file manipulation on Apollia OS.

Specialised agent that reads, analyses, creates and modifies Excel files
(.xlsx, .xlsm) using openpyxl.  Works on any LLM backend (7B and above).

Required packages (installed automatically at agent initialisation):
  openpyxl>=3.1.0
"""

from __future__ import annotations

from typing import Any

from apollia.agents import AIPResult, WorkerAgent

SYSTEM_PROMPT: str = """Tu es excel-worker, un agent expert de la manipulation de fichiers Excel via Python.

## RÈGLES ABSOLUES (non-négociables)

1. N'utilise JAMAIS bash_executor pour lire, écrire ou modifier un fichier .xlsx ou .xlsm.
   RAISON : Un fichier .xlsx est une archive ZIP. Toute manipulation bash corrompt l'archive silencieusement.

2. Utilise TOUJOURS openpyxl pour lire et écrire des fichiers .xlsx/.xlsm.
   Pour les fichiers .xls (ancien format Binary BIFF) : utilise xlrd en lecture seule
   et informe l'utilisateur que l'édition n'est pas possible sans conversion préalable.

3. TOUJOURS appeler wb.save(path) après toute modification.
   Sans wb.save(), les changements restent uniquement en mémoire — le fichier disque reste inchangé.

4. Ne jamais supposer le nom ou l'index d'une feuille.
   Toujours inspecter wb.sheetnames avant d'accéder à une feuille spécifique.

## IMPORTS STANDARDS

```python
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
```

## PATTERNS OBLIGATOIRES

### Lire un fichier existant
```python
wb = load_workbook(path, read_only=True, data_only=True)
sheets = wb.sheetnames
ws = wb[sheets[0]]              # ou wb["NomFeuille"] si précisé par l'utilisateur
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
wb.close()                      # Toujours fermer en mode read_only
```

### Modifier un fichier existant
```python
wb = load_workbook(path)        # JAMAIS read_only=True si on modifie
ws = wb.active                  # ou wb["NomFeuille"]
ws["A1"] = nouvelle_valeur
wb.save(path)                   # OBLIGATOIRE — sans ça les changements sont perdus
```

### Créer un nouveau fichier
```python
wb = Workbook()
ws = wb.active
ws.title = "Données"
ws.append(["Colonne1", "Colonne2", "Colonne3"])   # En-têtes
ws.append([valeur1, valeur2, valeur3])              # Données
wb.save(output_path)
```

## GESTION DES ERREURS DOMAINE

- `FileNotFoundError` → message d'erreur clair : "Fichier introuvable : {path}"
- `zipfile.BadZipFile` ou `InvalidFileException` → informer que le fichier est corrompu
- `KeyError` sur nom de feuille → lister les feuilles disponibles (wb.sheetnames)
- Cellule hors range → ne pas planter, retourner None silencieusement

## FORMAT DE RÉPONSE

- Toujours indiquer la feuille traitée et le nombre de lignes/colonnes trouvées
- Pour les analyses : présenter les résultats en tableau Markdown
- Pour les modifications : confirmer le nombre de cellules modifiées + le path sauvegardé
- Pour les erreurs : message clair avec le path du fichier et la raison précise
"""


def manifest() -> dict[str, Any]:
    """Return the AIP agent manifest for excel-worker."""
    return {
        "name": "excel-worker",
        "version": "0.1.0",
        "description": (
            "Agent spécialisé pour la manipulation de fichiers Excel (.xlsx, .xlsm). "
            "Lit, analyse, crée et modifie des classeurs Excel via openpyxl. "
            "Fonctionne sur tous les modèles LLM (7B+)."
        ),
        "execution_mode": "direct",
        "tools_required": ["python_executor", "file_read", "file_write"],
        "tools_optional": ["file_list"],
        "tools_requiring_approval": [],
        "packages": ["openpyxl>=3.1.0"],
        "memory_namespace": "excel-worker",
        "supports_a2a": True,
        "skills": [
            {
                "id": "read-excel",
                "name": "Lire un fichier Excel",
                "description": "Lit et retourne le contenu structuré d'une feuille Excel (headers + rows)",
                "input_modes": ["text"],
                "output_modes": ["text"],
            },
            {
                "id": "edit-excel",
                "name": "Modifier un fichier Excel",
                "description": "Modifie des cellules, ajoute des lignes ou colonnes",
                "input_modes": ["text"],
                "output_modes": ["text"],
            },
            {
                "id": "analyze-excel",
                "name": "Analyser les données Excel",
                "description": "Calcule des statistiques : totaux, moyennes, recherche par valeur",
                "input_modes": ["text"],
                "output_modes": ["text"],
            },
        ],
        "tags": ["excel", "spreadsheet", "xlsx", "data", "worker"],
        "max_concurrent_tasks": 1,
        "dangerous_tools_allowed": False,
    }


class ExcelWorkerAgent(WorkerAgent):
    """Worker Agent specialised for Excel file manipulation via openpyxl.

    The expertise (correct API patterns, guardrails, error handling) is
    compiled into SYSTEM_PROMPT and the class, not injected at runtime.
    This makes the agent reliable on small LLMs (7B+) that cannot reliably
    discover openpyxl patterns on their own.
    """

    SYSTEM_PROMPT = SYSTEM_PROMPT
    MAX_STEPS = 8
    TEMPERATURE = 0.1

    def manifest(self) -> dict[str, Any]:
        """Return the AIP agent manifest for excel-worker."""
        return manifest()

    async def run(self, task: dict[str, Any], ctx: Any) -> dict[str, Any]:
        """Execute the Excel task using the ReAct loop.

        Extracts the user message from the task input, delegates to the
        inherited ReAct loop, and wraps the result in an AIPResult dict.
        """
        user_message = (
            task.get("input", {}).get("text", "")
            if isinstance(task.get("input"), dict)
            else str(task.get("input", ""))
        )
        result = await self.react(task, ctx, user_message)
        if isinstance(result, dict):
            return result
        return AIPResult.completed(result)


agent = ExcelWorkerAgent()
