"""xlsx-worker — Manipulation de fichiers Excel (.xlsx).

Worker Apollia OS standalone qui isole la manipulation .xlsx du LLM director.
Cinq skills A2A déterministes :

- ``xlsx.read``              — lecture cellules brutes, types, formules, merged ranges.
- ``xlsx.read-as-dataframe`` — lecture via pandas (dtype inference + header detection).
- ``xlsx.write``             — écriture multi-sheet avec styles avancés et conditional formatting.
- ``xlsx.append-rows``       — ajout de lignes sans toucher au reste.
- ``xlsx.update-cells``      — patch ciblé par A1 notation.

Dispatch multi-skills : le runtime Apollia (≥ 2026-05-19) propage le
``skill_id`` invoqué dans ``AIPTask``. Le worker lit ``task.skill_id`` via
``apollia.utils.a2a.extract_skill_id(task)`` et dispatche sur le full
identifier (``xlsx.read``, ``xlsx.write``, …).
"""

from __future__ import annotations

import datetime
import json
from pathlib import Path
from typing import Any

from apollia import DomainError, agent, skill
from apollia.types import Ctx

MAX_FILE_BYTES: int = 100 * 1024 * 1024     # 100 MiB
EXCEL_ROW_LIMIT: int = 1_048_576            # Excel hard limit

CF_RULES: tuple[str, ...] = (
    "greater_than",
    "less_than",
    "equal_to",
    "not_equal_to",
    "between",
    "not_between",
    "contains_text",
    "not_contains_text",
    "begins_with",
    "ends_with",
)

DTYPE_MAP: dict[str, str] = {
    "int": "Int64",
    "float": "float64",
    "str": "string",
    "bool": "boolean",
    "datetime": "datetime64[ns]",
}


class _XlsxError(DomainError):
    """Legacy alias — subclass of :class:`DomainError` for typed dispatch."""


@agent(
    name="xlsx-worker",
    version="0.1.0",
    description=(
        "Manipulation de fichiers Excel (.xlsx) : lire, écrire, ajouter "
        "des lignes, patcher des cellules ciblées — préserve types, "
        "formules et merged cells. Styles avancés et lecture pandas."
    ),
    tags=("xlsx", "excel", "spreadsheet", "openpyxl", "pandas", "worker"),
    agent_type="worker",
    step_budget={"max_steps": 1, "max_tool_calls": 5, "wall_clock_secs": 300},
)
class XlsxWorker:
    """5 skills .xlsx déterministes (decorator-first)."""

    @skill(
        "xlsx.read",
        description=(
            "Lit un fichier .xlsx et retourne JSON structuré (feuilles, "
            "lignes typées, merged ranges, freeze panes). Mode 'values' ou 'formulas'."
        ),
    )
    async def read_xlsx(
        self,
        path: str,
        sheet_names: list[str] | None = None,
        read_mode: str = "values",
        max_rows: int = 100_000,
        max_cols: int = 1_000,
        include_types: bool = False,
        ctx: Ctx = None,
    ) -> dict[str, Any]:
        """Lecture brute."""
        return _do_read(
            {
                "path": path,
                "sheet_names": sheet_names,
                "read_mode": read_mode,
                "max_rows": max_rows,
                "max_cols": max_cols,
                "include_types": include_types,
            }
        )

    @skill(
        "xlsx.read_as_dataframe",
        description=(
            "Lit un .xlsx en pandas.DataFrame, infère les dtypes, retourne "
            "les records JSON."
        ),
    )
    async def read_as_dataframe(
        self,
        path: str,
        sheet_name: str | int | None = None,
        header_row: int | None = 0,
        skiprows: int = 0,
        nrows: int | None = None,
        dtypes_hint: dict[str, str] | None = None,
        na_values: list[Any] | None = None,
        ctx: Ctx = None,
    ) -> dict[str, Any]:
        """Lecture pandas."""
        return _do_read_as_dataframe(
            {
                "path": path,
                "sheet_name": sheet_name,
                "header_row": header_row,
                "skiprows": skiprows,
                "nrows": nrows,
                "dtypes_hint": dtypes_hint,
                "na_values": na_values,
            }
        )

    @skill(
        "xlsx.write",
        description=(
            "Crée un .xlsx multi-sheet avec styles avancés, conditional "
            "formatting, freeze panes, auto-filter."
        ),
    )
    async def write_xlsx(
        self,
        output_path: str,
        sheets: list[dict[str, Any]],
        named_styles: dict[str, Any] | None = None,
        overwrite: bool = False,
        ctx: Ctx = None,
    ) -> dict[str, Any]:
        """Écriture multi-sheet."""
        return _do_write(
            {
                "output_path": output_path,
                "sheets": sheets,
                "named_styles": named_styles,
                "overwrite": overwrite,
            }
        )

    @skill(
        "xlsx.append_rows",
        description=(
            "Ajoute des lignes à une feuille existante sans toucher au reste."
        ),
    )
    async def append_rows(
        self,
        path: str,
        rows: list[list[Any]],
        sheet_name: str | None = None,
        ctx: Ctx = None,
    ) -> dict[str, Any]:
        """Ajout de lignes."""
        return _do_append_rows({"path": path, "rows": rows, "sheet_name": sheet_name})

    @skill(
        "xlsx.update_cells",
        description=(
            "Patch ciblé cellule par cellule (A1 notation). Préserve les "
            "formules des cellules non touchées."
        ),
    )
    async def update_cells(
        self,
        path: str,
        sheet_name: str,
        updates: list[dict[str, Any]],
        ctx: Ctx = None,
    ) -> dict[str, Any]:
        """Patch ciblé."""
        return _do_update_cells(
            {"path": path, "sheet_name": sheet_name, "updates": updates}
        )


# ─── Operations ────────────────────────────────────────────────────────────


def _do_read(payload: dict[str, Any]) -> dict[str, Any]:
    path = _require_str(payload, "path")
    _validate_xlsx_path(path, must_exist=True)

    sheet_names_filter = payload.get("sheet_names")
    read_mode = payload.get("read_mode", "values")
    max_rows = int(payload.get("max_rows", 100_000))
    max_cols = int(payload.get("max_cols", 1000))
    include_types = bool(payload.get("include_types", False))

    if read_mode not in ("values", "formulas"):
        raise _XlsxError(
            "INVALID_TYPE",
            f"read_mode doit être 'values' ou 'formulas'. Reçu : {read_mode!r}",
            details={"field": "read_mode"},
        )

    import openpyxl
    from openpyxl.utils import column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string

    try:
        wb = openpyxl.load_workbook(path, data_only=(read_mode == "values"), read_only=False)
    except Exception as exc:
        raise _XlsxError("PARSE_ERROR", f"Impossible de lire le .xlsx : {exc}") from exc

    target_names = sheet_names_filter if sheet_names_filter else wb.sheetnames
    sheets_out: list[dict[str, Any]] = []

    for name in target_names:
        if name not in wb.sheetnames:
            wb.close()
            raise _XlsxError(
                "SHEET_NOT_FOUND",
                f"Feuille introuvable : {name}",
                details={"sheet_name": name, "available": wb.sheetnames},
            )
        ws = wb[name]

        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0
        capped_rows = min(total_rows, max_rows)
        capped_cols = min(total_cols, max_cols)

        rows: list[list[Any]] = []
        for r in range(1, capped_rows + 1):
            row_data: list[Any] = []
            for c in range(1, capped_cols + 1):
                cell = ws.cell(row=r, column=c)
                value = _cell_value(cell, read_mode)
                if include_types:
                    row_data.append({"value": value, "type": _cell_type_name(cell, read_mode, value)})
                else:
                    row_data.append(value)
            rows.append(row_data)

        # Headers : première ligne si toutes les valeurs sont des strings
        headers: list[str] | None = None
        if rows:
            first_row_raw = rows[0]
            first_values = [c["value"] if isinstance(c, dict) else c for c in first_row_raw]
            if all(isinstance(v, str) and v for v in first_values):
                headers = first_values  # type: ignore[assignment]

        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        frozen_rows, frozen_cols = _parse_freeze_panes(
            ws.freeze_panes, coordinate_from_string, column_index_from_string
        )

        sheets_out.append({
            "name": name,
            "headers": headers,
            "rows": rows,
            "total_rows": total_rows,
            "total_cols": total_cols,
            "merged_ranges": merged_ranges,
            "frozen_rows": frozen_rows,
            "frozen_cols": frozen_cols,
            "truncated": total_rows > max_rows or total_cols > max_cols,
        })

    wb.close()
    return {"path": path, "read_mode": read_mode, "sheets": sheets_out}


def _do_read_as_dataframe(payload: dict[str, Any]) -> dict[str, Any]:
    path = _require_str(payload, "path")
    _validate_xlsx_path(path, must_exist=True)

    sheet_name = payload.get("sheet_name")
    header_row = payload.get("header_row", 0)
    skiprows = payload.get("skiprows")
    nrows = payload.get("nrows")
    dtypes_hint = payload.get("dtypes_hint")
    na_values = payload.get("na_values")

    pd_kwargs: dict[str, Any] = {"io": path, "engine": "openpyxl"}
    if sheet_name is not None:
        pd_kwargs["sheet_name"] = sheet_name
    if header_row is None:
        pd_kwargs["header"] = None
    else:
        pd_kwargs["header"] = int(header_row)
    if skiprows is not None:
        pd_kwargs["skiprows"] = int(skiprows)
    if nrows is not None:
        pd_kwargs["nrows"] = int(nrows)
    if na_values:
        pd_kwargs["na_values"] = na_values

    date_cols: list[str] = []
    if dtypes_hint:
        if not isinstance(dtypes_hint, dict):
            raise _XlsxError(
                "INVALID_TYPE",
                "dtypes_hint doit être un objet {col: type_str}",
                details={"field": "dtypes_hint"},
            )
        pd_dtypes: dict[str, str] = {}
        for col, dtype_str in dtypes_hint.items():
            if dtype_str not in DTYPE_MAP:
                raise _XlsxError(
                    "INVALID_TYPE",
                    f"dtypes_hint['{col}'] invalide : {dtype_str!r}. Valeurs : {list(DTYPE_MAP)}",
                    details={"field": "dtypes_hint", "got": dtype_str},
                )
            if dtype_str == "datetime":
                date_cols.append(col)
            else:
                pd_dtypes[col] = DTYPE_MAP[dtype_str]
        if pd_dtypes:
            pd_kwargs["dtype"] = pd_dtypes
        if date_cols:
            pd_kwargs["parse_dates"] = date_cols

    import pandas as pd

    try:
        df = pd.read_excel(**pd_kwargs)
    except FileNotFoundError as exc:
        raise _XlsxError("FILE_NOT_FOUND", str(exc)) from exc
    except Exception as exc:
        msg = str(exc).lower()
        if "could not convert" in msg or "could not cast" in msg or "invalid literal" in msg:
            raise _XlsxError("DTYPE_COERCION_FAILED", str(exc), details={"hint": dtypes_hint}) from exc
        raise _XlsxError("PARSE_ERROR", f"Impossible de lire le .xlsx : {exc}") from exc

    na_count = int(df.isna().sum().sum())
    # Utilise to_json + json.loads pour sérialiser proprement (gère Timestamp, numpy types, NaN).
    records = json.loads(df.to_json(orient="records", date_format="iso", default_handler=str))
    dtypes_out = {str(col): str(df.dtypes[col]) for col in df.columns}

    effective_sheet = sheet_name if sheet_name is not None else _first_sheet_name(path)

    return {
        "path": path,
        "sheet_name": effective_sheet,
        "columns": [str(c) for c in df.columns],
        "dtypes": dtypes_out,
        "records": records,
        "rows_count": int(len(df)),
        "na_count": na_count,
    }


def _do_write(payload: dict[str, Any]) -> dict[str, Any]:
    output_path = _require_str(payload, "output_path")
    _validate_xlsx_extension(output_path)
    overwrite = bool(payload.get("overwrite", False))

    out = Path(output_path).expanduser().resolve()
    if out.exists() and not overwrite:
        raise _XlsxError(
            "OUTPUT_EXISTS",
            f"Fichier de sortie existe déjà (overwrite=false) : {out}",
            details={"output_path": str(out)},
        )
    out.parent.mkdir(parents=True, exist_ok=True)

    named_styles_raw = payload.get("named_styles") or {}
    sheets_spec = payload.get("sheets")
    if not isinstance(sheets_spec, list) or not sheets_spec:
        raise _XlsxError(
            "MISSING_FIELD",
            "Champ 'sheets' requis (liste non vide)",
            details={"field": "sheets"},
        )

    style_objects: dict[str, dict[str, Any]] = {}
    for style_id, spec in named_styles_raw.items():
        try:
            style_objects[style_id] = _build_style_dict(spec)
        except _XlsxError:
            raise
        except Exception as exc:
            raise _XlsxError(
                "INVALID_STYLE",
                f"Style '{style_id}' invalide : {exc}",
                details={"style_id": style_id},
            ) from exc

    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet

    styles_applied_count = 0
    conditional_rules_count = 0
    total_rows_written = 0
    total_cells_written = 0

    for sheet_idx, sheet_spec in enumerate(sheets_spec):
        if not isinstance(sheet_spec, dict):
            raise _XlsxError(
                "INVALID_TYPE",
                f"sheets[{sheet_idx}] doit être un objet",
                details={"field": f"sheets[{sheet_idx}]"},
            )
        name = sheet_spec.get("name")
        if not isinstance(name, str) or not name:
            raise _XlsxError(
                "MISSING_FIELD",
                f"sheets[{sheet_idx}].name requis (string non vide)",
                details={"field": f"sheets[{sheet_idx}].name"},
            )

        ws = wb.create_sheet(name)

        rows_input = sheet_spec.get("rows")
        if not isinstance(rows_input, list):
            raise _XlsxError(
                "MISSING_FIELD",
                f"sheets[{sheet_idx}].rows requis (liste)",
                details={"field": f"sheets[{sheet_idx}].rows"},
            )
        if len(rows_input) > EXCEL_ROW_LIMIT:
            raise _XlsxError(
                "TOO_LARGE",
                f"Trop de lignes ({len(rows_input)} > {EXCEL_ROW_LIMIT})",
                details={"row_count": len(rows_input), "limit": EXCEL_ROW_LIMIT},
            )

        headers = sheet_spec.get("headers")
        next_row = 1
        if headers:
            if not isinstance(headers, list):
                raise _XlsxError(
                    "INVALID_TYPE",
                    f"sheets[{sheet_idx}].headers doit être une liste",
                    details={"field": f"sheets[{sheet_idx}].headers"},
                )
            for c_idx, h in enumerate(headers, start=1):
                ws.cell(row=1, column=c_idx, value=h)
            total_cells_written += len(headers)
            next_row = 2

        for r_offset, row in enumerate(rows_input):
            if not isinstance(row, list):
                raise _XlsxError(
                    "INVALID_TYPE",
                    f"sheets[{sheet_idx}].rows[{r_offset}] doit être une liste",
                    details={"field": f"sheets[{sheet_idx}].rows[{r_offset}]"},
                )
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=next_row + r_offset, column=c_idx, value=_coerce_write_value(val))
                total_cells_written += 1
            total_rows_written += 1

        # Column widths
        for col_letter, width in (sheet_spec.get("column_widths") or {}).items():
            try:
                ws.column_dimensions[str(col_letter).upper()].width = float(width)
            except (ValueError, TypeError) as exc:
                raise _XlsxError(
                    "INVALID_TYPE",
                    f"column_widths['{col_letter}'] invalide : {exc}",
                    details={"col_letter": col_letter, "width": width},
                ) from exc

        # Row heights
        for row_str, height in (sheet_spec.get("row_heights") or {}).items():
            try:
                ws.row_dimensions[int(row_str)].height = float(height)
            except (ValueError, TypeError) as exc:
                raise _XlsxError(
                    "INVALID_TYPE",
                    f"row_heights['{row_str}'] invalide : {exc}",
                    details={"row": row_str, "height": height},
                ) from exc

        # Freeze panes
        freeze = sheet_spec.get("freeze") or {}
        f_row = int(freeze.get("row", 0))
        f_col = int(freeze.get("col", 0))
        if f_row > 0 or f_col > 0:
            col_letter = get_column_letter(f_col + 1)
            ws.freeze_panes = f"{col_letter}{f_row + 1}"

        # Auto-filter
        if sheet_spec.get("auto_filter") and ws.dimensions:
            ws.auto_filter.ref = ws.dimensions

        # Styles apply
        for sa in (sheet_spec.get("styles_apply") or []):
            style_id = sa.get("style")
            if style_id not in style_objects:
                raise _XlsxError(
                    "STYLE_NOT_FOUND",
                    f"Style '{style_id}' référencé mais absent de named_styles",
                    details={"style_id": style_id, "available": list(style_objects)},
                )
            target_range = sa.get("range") or sa.get("cell_ref")
            if not target_range:
                raise _XlsxError(
                    "MISSING_FIELD",
                    "styles_apply entry requires 'range' or 'cell_ref'",
                    details={"field": "styles_apply"},
                )
            styles_applied_count += _apply_style_to_range(ws, str(target_range), style_objects[style_id])

        # Conditional formatting
        for cf in (sheet_spec.get("conditional_formatting") or []):
            _add_conditional_format(ws, cf, style_objects)
            conditional_rules_count += 1

    try:
        wb.save(out)
    except Exception as exc:
        raise _XlsxError("EXECUTION_FAILED", f"Sauvegarde impossible : {exc}") from exc
    wb.close()

    return {
        "output_path": str(out),
        "sheets_count": len(sheets_spec),
        "total_rows_written": total_rows_written,
        "total_cells_written": total_cells_written,
        "styles_applied_count": styles_applied_count,
        "conditional_rules_count": conditional_rules_count,
    }


def _do_append_rows(payload: dict[str, Any]) -> dict[str, Any]:
    path = _require_str(payload, "path")
    _validate_xlsx_path(path, must_exist=True)

    rows_input = payload.get("rows")
    if not isinstance(rows_input, list) or not rows_input:
        raise _XlsxError(
            "MISSING_FIELD",
            "Champ 'rows' requis (liste non vide)",
            details={"field": "rows"},
        )

    sheet_name = payload.get("sheet_name")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=False, keep_vba=False)
    except Exception as exc:
        raise _XlsxError("PARSE_ERROR", f"Impossible de lire le .xlsx : {exc}") from exc

    if sheet_name is None:
        sheet_name = wb.sheetnames[0] if wb.sheetnames else None
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise _XlsxError(
            "SHEET_NOT_FOUND",
            f"Feuille introuvable : {sheet_name}",
            details={"sheet_name": sheet_name, "available": wb.sheetnames},
        )

    ws = wb[sheet_name]
    current_max = ws.max_row or 0
    if current_max + len(rows_input) > EXCEL_ROW_LIMIT:
        wb.close()
        raise _XlsxError(
            "TOO_LARGE",
            f"Ajout dépasserait la limite Excel ({current_max + len(rows_input)} > {EXCEL_ROW_LIMIT})",
            details={
                "current": current_max,
                "appending": len(rows_input),
                "limit": EXCEL_ROW_LIMIT,
            },
        )

    appended = 0
    for idx, row in enumerate(rows_input):
        if not isinstance(row, list):
            wb.close()
            raise _XlsxError(
                "INVALID_TYPE",
                f"rows[{idx}] doit être une liste",
                details={"index": idx},
            )
        ws.append([_coerce_write_value(v) for v in row])
        appended += 1

    try:
        wb.save(path)
    except Exception as exc:
        raise _XlsxError("EXECUTION_FAILED", f"Sauvegarde impossible : {exc}") from exc

    new_max = ws.max_row
    wb.close()

    return {
        "path": path,
        "sheet_name": sheet_name,
        "rows_appended": appended,
        "new_total_rows": new_max,
    }


def _do_update_cells(payload: dict[str, Any]) -> dict[str, Any]:
    path = _require_str(payload, "path")
    _validate_xlsx_path(path, must_exist=True)
    sheet_name = _require_str(payload, "sheet_name")

    updates = payload.get("updates")
    if not isinstance(updates, list) or not updates:
        raise _XlsxError(
            "MISSING_FIELD",
            "Champ 'updates' requis (liste non vide)",
            details={"field": "updates"},
        )

    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=False, keep_vba=False)
    except Exception as exc:
        raise _XlsxError("PARSE_ERROR", f"Impossible de lire le .xlsx : {exc}") from exc

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise _XlsxError(
            "SHEET_NOT_FOUND",
            f"Feuille introuvable : {sheet_name}",
            details={"sheet_name": sheet_name, "available": wb.sheetnames},
        )

    ws = wb[sheet_name]
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for idx, upd in enumerate(updates):
        if not isinstance(upd, dict):
            errors.append({"index": idx, "reason": "update entry must be an object"})
            skipped += 1
            continue
        cell_ref = upd.get("cell_ref")
        if not isinstance(cell_ref, str) or not cell_ref:
            errors.append({"index": idx, "reason": "missing or invalid cell_ref"})
            skipped += 1
            continue
        try:
            ws[cell_ref] = _coerce_write_value(upd.get("value"))
            updated += 1
        except Exception as exc:
            errors.append({"cell_ref": cell_ref, "reason": str(exc)})
            skipped += 1

    try:
        wb.save(path)
    except Exception as exc:
        raise _XlsxError("EXECUTION_FAILED", f"Sauvegarde impossible : {exc}") from exc
    wb.close()

    return {
        "path": path,
        "sheet_name": sheet_name,
        "updated_count": updated,
        "skipped_count": skipped,
        "errors": errors,
    }


# ─── Style building ────────────────────────────────────────────────────────


def _build_style_dict(spec: Any) -> dict[str, Any]:
    """Convert a StyleSpec dict to a dict of openpyxl style objects."""
    if not isinstance(spec, dict):
        raise _XlsxError("INVALID_STYLE", "StyleSpec doit être un objet")

    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    style: dict[str, Any] = {}

    font_spec = spec.get("font")
    if font_spec:
        if not isinstance(font_spec, dict):
            raise _XlsxError("INVALID_STYLE", "font doit être un objet")
        color = font_spec.get("color")
        style["font"] = Font(
            name=font_spec.get("name"),
            size=font_spec.get("size"),
            bold=font_spec.get("bold"),
            italic=font_spec.get("italic"),
            underline="single" if font_spec.get("underline") else None,
            strike=font_spec.get("strike"),
            color=_validate_hex_color(color) if color else None,
        )

    fill_spec = spec.get("fill")
    if fill_spec:
        if not isinstance(fill_spec, dict):
            raise _XlsxError("INVALID_STYLE", "fill doit être un objet")
        color = fill_spec.get("color")
        if color:
            style["fill"] = PatternFill(
                fgColor=_validate_hex_color(color),
                patternType=fill_spec.get("pattern", "solid"),
            )

    border_spec = spec.get("border")
    if border_spec:
        if not isinstance(border_spec, dict):
            raise _XlsxError("INVALID_STYLE", "border doit être un objet")

        def _side(s: Any) -> Side | None:
            if not s:
                return None
            if not isinstance(s, dict):
                raise _XlsxError("INVALID_STYLE", "border side doit être un objet")
            return Side(
                style=s.get("style"),
                color=_validate_hex_color(s.get("color")) if s.get("color") else None,
            )

        style["border"] = Border(
            top=_side(border_spec.get("top")),
            bottom=_side(border_spec.get("bottom")),
            left=_side(border_spec.get("left")),
            right=_side(border_spec.get("right")),
        )

    align_spec = spec.get("alignment")
    if align_spec:
        if not isinstance(align_spec, dict):
            raise _XlsxError("INVALID_STYLE", "alignment doit être un objet")
        style["alignment"] = Alignment(
            horizontal=align_spec.get("horizontal"),
            vertical=align_spec.get("vertical"),
            wrap_text=align_spec.get("wrap_text"),
            indent=align_spec.get("indent", 0),
        )

    nf = spec.get("number_format")
    if nf is not None:
        if not isinstance(nf, str):
            raise _XlsxError("INVALID_STYLE", "number_format doit être string")
        style["number_format"] = nf

    return style


def _validate_hex_color(color: Any) -> str:
    """Accept '#RRGGBB' or 'RRGGBB' (case-insensitive). Return 'FFRRGGBB' (ARGB)."""
    if not isinstance(color, str):
        raise _XlsxError(
            "INVALID_STYLE",
            f"Couleur doit être hex string, reçu : {type(color).__name__}",
        )
    c = color.lstrip("#").upper()
    if len(c) == 6 and all(ch in "0123456789ABCDEF" for ch in c):
        return "FF" + c
    if len(c) == 8 and all(ch in "0123456789ABCDEF" for ch in c):
        return c
    raise _XlsxError(
        "INVALID_STYLE",
        f"Couleur hex invalide : {color!r} (attendu RRGGBB ou AARRGGBB)",
    )


def _apply_style_to_range(ws: Any, range_str: str, style_dict: dict[str, Any]) -> int:
    """Apply a style dict to every cell in a range. Returns the cell count."""
    min_col, min_row, max_col, max_row = _parse_range(ws, range_str)
    count = 0
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if "font" in style_dict:
                cell.font = style_dict["font"]
            if "fill" in style_dict:
                cell.fill = style_dict["fill"]
            if "border" in style_dict:
                cell.border = style_dict["border"]
            if "alignment" in style_dict:
                cell.alignment = style_dict["alignment"]
            if "number_format" in style_dict:
                cell.number_format = style_dict["number_format"]
            count += 1
    return count


def _parse_range(ws: Any, range_str: str) -> tuple[int, int, int, int]:
    """Parse 'A1', 'A1:C3', 'A:C', '1:3'. Returns (min_col, min_row, max_col, max_row)."""
    from openpyxl.utils.cell import range_boundaries, column_index_from_string

    parts = range_str.split(":")
    # Full row range "1:3"
    if len(parts) == 2 and all(p.isdigit() for p in parts):
        r1, r2 = int(parts[0]), int(parts[1])
        return 1, min(r1, r2), ws.max_column or 1, max(r1, r2)
    # Full column range "A:C"
    if len(parts) == 2 and all(p.isalpha() for p in parts):
        c1 = column_index_from_string(parts[0].upper())
        c2 = column_index_from_string(parts[1].upper())
        return min(c1, c2), 1, max(c1, c2), ws.max_row or 1
    # Single cell or normal range
    target = range_str if ":" in range_str else f"{range_str}:{range_str}"
    try:
        return range_boundaries(target)
    except Exception as exc:
        raise _XlsxError(
            "INVALID_RANGE",
            f"Range invalide : {range_str!r} ({exc})",
            details={"range": range_str},
        ) from exc


def _add_conditional_format(ws: Any, cf_spec: Any, style_objects: dict[str, dict[str, Any]]) -> None:
    if not isinstance(cf_spec, dict):
        raise _XlsxError("INVALID_RULE", "conditional_formatting entry doit être un objet")

    target_range = cf_spec.get("range")
    rule_name = cf_spec.get("rule")
    style_id = cf_spec.get("style")

    if not target_range:
        raise _XlsxError(
            "MISSING_FIELD",
            "conditional_formatting needs 'range'",
            details={"field": "conditional_formatting.range"},
        )
    if not rule_name:
        raise _XlsxError(
            "MISSING_FIELD",
            "conditional_formatting needs 'rule'",
            details={"field": "conditional_formatting.rule"},
        )
    if rule_name not in CF_RULES:
        raise _XlsxError(
            "INVALID_RULE",
            f"Rule '{rule_name}' non supportée. Valeurs : {list(CF_RULES)}",
            details={"rule": rule_name},
        )
    if style_id not in style_objects:
        raise _XlsxError(
            "STYLE_NOT_FOUND",
            f"Style '{style_id}' référencé en conditional_formatting absent de named_styles",
            details={"style_id": style_id, "available": list(style_objects)},
        )

    from openpyxl.formatting.rule import CellIsRule, FormulaRule

    style_dict = style_objects[style_id]
    common_kwargs = {
        "stopIfTrue": False,
        "fill": style_dict.get("fill"),
        "font": style_dict.get("font"),
        "border": style_dict.get("border"),
    }

    if rule_name in ("greater_than", "less_than", "equal_to", "not_equal_to"):
        op_map = {
            "greater_than": "greaterThan",
            "less_than": "lessThan",
            "equal_to": "equal",
            "not_equal_to": "notEqual",
        }
        value = cf_spec.get("value")
        if value is None:
            raise _XlsxError(
                "MISSING_FIELD",
                f"Rule '{rule_name}' requires 'value'",
                details={"rule": rule_name},
            )
        rule = CellIsRule(operator=op_map[rule_name], formula=[str(value)], **common_kwargs)
    elif rule_name in ("between", "not_between"):
        operator = "between" if rule_name == "between" else "notBetween"
        values = cf_spec.get("values")
        if not (isinstance(values, list) and len(values) == 2):
            raise _XlsxError(
                "MISSING_FIELD",
                f"Rule '{rule_name}' requires 'values' as [a, b]",
                details={"rule": rule_name},
            )
        rule = CellIsRule(
            operator=operator,
            formula=[str(values[0]), str(values[1])],
            **common_kwargs,
        )
    else:  # text-based rules
        text_val = cf_spec.get("value")
        if not isinstance(text_val, str):
            raise _XlsxError(
                "MISSING_FIELD",
                f"Rule '{rule_name}' requires 'value' as string",
                details={"rule": rule_name},
            )
        first_cell = target_range.split(":")[0] if ":" in target_range else target_range
        text_escaped = text_val.replace('"', '""')
        if rule_name == "contains_text":
            formula = f'NOT(ISERROR(SEARCH("{text_escaped}",{first_cell})))'
        elif rule_name == "not_contains_text":
            formula = f'ISERROR(SEARCH("{text_escaped}",{first_cell}))'
        elif rule_name == "begins_with":
            formula = f'LEFT({first_cell},{len(text_val)})="{text_escaped}"'
        else:  # ends_with
            formula = f'RIGHT({first_cell},{len(text_val)})="{text_escaped}"'
        rule = FormulaRule(formula=[formula], **common_kwargs)

    ws.conditional_formatting.add(target_range, rule)


# ─── Helpers ───────────────────────────────────────────────────────────────


def _require_str(payload: dict[str, Any], field: str) -> str:
    value = payload.get(field)
    if not isinstance(value, str) or not value:
        raise _XlsxError(
            "MISSING_FIELD",
            f"Champ '{field}' requis (string non vide)",
            details={"field": field},
        )
    return value


def _validate_xlsx_extension(path: str) -> None:
    p = path.lower()
    if p.endswith(".xlsm") or p.endswith(".xls") or p.endswith(".xlsb"):
        raise _XlsxError(
            "UNSUPPORTED_FORMAT",
            f"Format non supporté : {path!r}. Uniquement .xlsx en v0.1.0.",
            details={"path": path},
        )
    if not p.endswith(".xlsx"):
        raise _XlsxError(
            "UNSUPPORTED_FORMAT",
            f"Extension non reconnue : {path!r}. .xlsx requis.",
            details={"path": path},
        )


def _validate_xlsx_path(path: str, must_exist: bool = False) -> None:
    _validate_xlsx_extension(path)
    if must_exist:
        p = Path(path)
        if not p.exists():
            raise _XlsxError("FILE_NOT_FOUND", f"fichier introuvable : {path}")
        size = p.stat().st_size
        if size > MAX_FILE_BYTES:
            raise _XlsxError(
                "TOO_LARGE",
                f"Fichier trop volumineux ({size} > {MAX_FILE_BYTES} bytes)",
                details={"size_bytes": size, "limit_bytes": MAX_FILE_BYTES},
            )


def _cell_value(cell: Any, read_mode: str) -> Any:
    if read_mode == "formulas" and cell.data_type == "f":
        return cell.value  # formula text e.g. "=SUM(A1:A10)"
    return cell.value


def _cell_type_name(cell: Any, read_mode: str, value: Any) -> str:
    if value is None:
        return "null"
    if read_mode == "formulas" and cell.data_type == "f":
        return "formula"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, datetime.datetime):
        return "datetime"
    if isinstance(value, datetime.date):
        return "date"
    if isinstance(value, str):
        return "string"
    if cell.data_type == "e":
        return "error"
    return "unknown"


def _parse_freeze_panes(
    freeze: Any,
    coordinate_from_string: Any,
    column_index_from_string: Any,
) -> tuple[int, int]:
    """Parse freeze panes cell ref (e.g., 'B2') → (frozen_rows, frozen_cols)."""
    if not freeze:
        return 0, 0
    try:
        col_str, row_int = coordinate_from_string(freeze)
        col_int = column_index_from_string(col_str)
        return row_int - 1, col_int - 1
    except Exception:
        return 0, 0


def _coerce_write_value(v: Any) -> Any:
    """Coerce a JSON-decoded value to an openpyxl-friendly Python value."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v
    # list / dict / autre → fallback string
    return str(v)


def _first_sheet_name(path: str) -> str | None:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = wb.sheetnames[0] if wb.sheetnames else None
    wb.close()
    return name


def _json_default(o: Any) -> Any:
    if isinstance(o, (datetime.datetime, datetime.date)):
        return o.isoformat()
    raise TypeError(f"Object of type {type(o).__name__} is not JSON serializable")


# (Module-level ``agent`` singleton exposé automatiquement par ``@agent``.)
